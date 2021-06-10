import datetime
import os
import main
import fake_useragent
import requests
from bs4 import BeautifulSoup
import openpyxl

t1 = datetime.datetime.now()
wbr_sogl = openpyxl.load_workbook("C://Users//user/Documents/документы/Взятые номера согласований.xlsx")
wbr_postan = openpyxl.load_workbook("C:\\Users\\user\Documents\документы\Номера постановлений.xlsx")
sh_sogl = wbr_sogl.active
sh_postan = wbr_postan.active
lrow_sogl = sh_sogl.max_row
lrow_postan = sh_postan.max_row
summ = lrow_sogl - lrow_postan
if lrow_sogl > lrow_postan:
    print('Имеются неотправленные постановления - ', summ)

    session = requests.Session()




    link = 'https://sed.rospotrebnadzor.ru/auth.php?group_id=3204'
    apply_link = 'https://sed.rospotrebnadzor.ru/document.php?all=1&c_user_status=0&dcv_status=1&category=6&DNSID=wXwaCw5PYykCpa7Zla8AjhQ'
    revision_link = 'https://sed.rospotrebnadzor.ru/document.php?all=1&c_user_status=1&category=6&DNSID=w9gfvomxjNRwo_HbkP_MOCg'
    registred_link = 'https://sed.rospotrebnadzor.ru/document.php?all=1&c_user_status=2&c_user_status_type=1&category=6&DNSID=w9gfvomxjNRwo_HbkP_MOCg'
    user = fake_useragent.UserAgent().random

    header = {
        'user-agent': user
    }

    data = {
        'DNSID': 'wqx7cahFJjk6x2QyA8VYkQg',
        'group_id': '32043',
        'user_id': '735221',
        'password': 'AlDmZh0510',
        'x': '1',

    }

    response = session.post(link, data=data, headers=header)


    responce_applied = session.post(apply_link, data=data, headers=header).text
    apply_soup = BeautifulSoup(responce_applied, 'lxml')
    apply_body = apply_soup.find('nobr')
    try:
        if apply_body.text == 'Нет записей':
            print('Все реально согласовано')
            n1 = 1
    except:
        print('Еще не все согласовано')
        n1 = 0



    responce_revision = session.post(revision_link, data=data, headers=header).text
    revision_soup = BeautifulSoup(responce_revision, 'lxml')
    revision_body = revision_soup.find('nobr')
    try:
        if revision_body.text == 'Нет записей':
            print('Возвратов реально нет')
            n2 = 1
    except:
        print('Что то вернули на доработку')
        n2 = 0

    responce_registred = session.post(registred_link, data=data, headers=header).text
    registred_soup = BeautifulSoup(responce_registred, 'lxml')
    registred_body = registred_soup.find('nobr')

    try:
        if registred_body.text == 'Нет записей':
            print('Все реально зарегистрировано')
            n3 = 1
    except:
        print('Еще не все зарегистрировано')
        n3 = 0

    if n1+n2+n3>2:
        print('Организуем отправку...')
        try:
            main.Otpravka.Sending()
        except Exception as ex:
            print(ex)
    if n1+n2+n3<3:
        print('Согласование или регистрация не завершена')

if lrow_sogl <= lrow_postan:
    print('Все что было зарегистрировано - отправлено! Больших нет...')
t2 = datetime.datetime.now()
t3 = t2 - t1
print(' ')
print('на все ушло', t3, 'секунд')
os.system('Taskkill /IM firefox.exe /F')