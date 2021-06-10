import locale

locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
import locale

locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
import glob
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl import Workbook
import datetime
import os
from selenium.webdriver.firefox.options import Options
from docxtpl import DocxTemplate
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from docx2pdf import convert

from openpyxl.styles import Color, PatternFill, Font, Border
def ARRIVED(filename):
    try:
        wbw = openpyxl.load_workbook("C:\\Users\\Администратор\\Documents" + "\\" + filename + ".xlsx")
        wbw = openpyxl.load_workbook("C:\\Users\\Администратор\\Documents" + "\\" + filename + ".xlsx")
        shr = wbw.active
        dlina = len(shr['E:E'])
        if shr["A1"].value is None:
            shr.cell(row=1, column=1, value='2')
        start = int(shr["A1"].value)



        options = Options()
        options.add_argument("--headless")
        options.set_preference("dom.webnotifications.enabled", False)
        browser = webdriver.Firefox(options=options)
        browser.get('https://52.gsen.ru/auth/?source_page=/?D9E18F1DA6F8841B&rand=i518010710')
        browser.find_element_by_id('idlogin').click()
        browser.find_element_by_id('idlogin').send_keys('tu52')
        browser.find_element_by_id('idpassword').click()
        browser.find_element_by_id('idpassword').send_keys('7393620')
        browser.find_element_by_xpath('//*[@id="idsubmit"]').click()
        browser.get('https://52.gsen.ru/rpn_int_planner/oper_person_via_epgu_list/?rand=i836061661')
        time.sleep(1)

        for n in range(start, dlina, 1):

            family = shr["E" + str(n)].value
            name = shr["F" + str(n)].value
            surname = shr["G" + str(n)].value

            print(family, " ", name, " ", surname)
            WebDriverWait(browser, 10).until(
                EC.element_to_be_clickable((By.NAME, 'search_fio_f')))
            browser.find_element_by_name('search_fio_f').click()
            browser.find_element_by_name('search_fio_f').send_keys(family)
            browser.find_element_by_name('search_fio_i').click()
            browser.find_element_by_name('search_fio_i').send_keys(name)
            browser.find_element_by_name('search_fio_o').click()
            if surname is not None:
                browser.find_element_by_name('search_fio_o').send_keys(surname)
            browser.find_element_by_class_name('submit.ui-button.ui-widget.ui-state-default.ui-corner-all').click()
            WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[1]/p[1]/b')))
            skolko = browser.find_element_by_xpath('/html/body/div[1]/div[1]/p[1]/b').text
            print(f'найдено {skolko} человек в базе')
            if int(skolko) < 1:
                print('Этот человек не зарегистрирован, будет запущена его регистрация и занесение в журнал')
                shr.cell(row=n, column=2, value='ОТСУТСТВУЕТ')
                wbw.save("C:\\Users\\Администратор\\Documents" + "\\" + filename + ".xlsx")
            if int(skolko) >=1:
                print('Этот человек зарегистрирован, будет процедура обработки...')
                shr.cell(row=n, column=2, value='ПРИСУТСТВУЕТ')
                wbw.save("C:\\Users\\Администратор\\Documents" + "\\" + filename + ".xlsx")
            browser.find_element_by_name('search_fio_f').clear()
            browser.find_element_by_name('search_fio_i').clear()
            browser.find_element_by_name('search_fio_o').clear()
            time.sleep(0.5)
            shr.cell(row=1, column=1, value=n)
        print('Операция по проверке завершена успешно! Всем спасибо, все свободны!')
        os.system('msg "user" Операция по проверке завершена успешно! Всем спасибо, все свободны!')
    except Exception as ex:
        print(ex)
        os.system(f'msg "user" Python закончил работу, убедитесь, что все хорошо {ex}')
    finally:
        os.system('Taskkill /IM firefox.exe /F')

def main():
    filename = input("Введите название проверяемого файла в папке ЗАГРУЗКИ...-")
    ARRIVED(filename)

if __name__ == "__main__":
    main()



